from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Create workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "Depo Malzeme Stok"

# Headers
headers = ["MALZEME EBAT (mm)", "KALINLIK (mm)", "BOY (MT)", "ADET", "ADET KG", "TOPLAM KG"]
ws.append(headers)

# Sample data
data = [
    ["250x250", 5, 6, 36, 233.89],
    ["250x250", 6, 6, 45, 276.03],
    ["250x250", 8, 6, 51, 368.24],
    ["250x250", 10, 6, 42, 471],
    ["200x300", 5, 6, 18, 230.84],
    ["200x300", 6, 6, 42, 257.71],
    ["200x300", 8, 6, 27, 360.92],
]

# Add data to worksheet
for row in data:
    ws.append(row)

# Excel formula to compute 'TOPLAM KG' based on geometry, not ADET KG
for i in range(2, 2 + len(data)):
    ws[f"F{i}"] = (
        f'=LET('
        f'enboy,TEXTSPLIT(SUBSTITUTE(A{i},"x","X"),"X"),'
        f'genislik,VALUE(TRIM(INDEX(enboy,1))),'
        f'yukseklik,VALUE(TRIM(INDEX(enboy,2))),'
        f'kalinlik,B{i},'
        f'boy_mm,C{i}*1000,'
        f'adet,D{i},'
        f'dis_alan,genislik*yukseklik,'
        f'ic_alan,(genislik-2*kalinlik)*(yukseklik-2*kalinlik),'
        f'kesit_alan,dis_alan-ic_alan,'
        f'hacim_mm3,kesit_alan*boy_mm,'
        f'agirlik_kg,(hacim_mm3*7.85)/1000000,'
        f'agirlik_kg*adet)'
    )

# Save file
file_path = "../../Desktop/Depo_Malzeme_Stok_Hesabi.xlsx"
wb.save(file_path)

print("Dosya başarıyla kaydedildi:", file_path)
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Create workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "Profil Ağırlık Hesabı"

# Headers
headers = ["EBAT (mm)", "KALINLIK (mm)", "BOY (MT)", "ADET", "Kilo (KG)"]
ws.append(headers)

# Sample data (rows)
data = [
    ["200x200", 8, 6, 1],
    ["250x250", 8, 6, 19],
    ["200x300", 10, 6, 42],
]

for row in data:
    ws.append(row)

# Add explanation/formula note to the header cell of the Kilo column
ws["F1"].comment = None
ws["E2"] = '=LET(enboy,TEXTSPLIT(SUBSTITUTE(A2,"x","X"),"X"),genislik,VALUE(TRIM(INDEX(enboy,1))),yukseklik,VALUE(TRIM(INDEX(enboy,2))),kalinlik,B2,boy_mm,C2*1000,adet,D2,dis_alan,genislik*yukseklik,ic_alan,(genislik-2*kalinlik)*(yukseklik-2*kalinlik),kesit_alan,dis_alan-ic_alan,hacim_mm3,kesit_alan*boy_mm,agirlik_kg,(hacim_mm3*7.85)/1000000,agirlik_kg*adet)'

# Save file
file_path = "../../Desktop/Profil_Agirlik_Hesabi.xlsx"
wb.save(file_path)

file_path